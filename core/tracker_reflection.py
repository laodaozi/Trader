"""
modules/tracker.py — 信号跟踪 · 绩效反思 · 标的新增

核心理念：
  不套用 Sharpe/Sortino/信息比率等标准框架。
  用自己的交易逻辑评价信号质量：
  
  信号发出 → 次日开盘入场 → 持有 N 天 → 判断这笔交易"对不对"

判定规则（自建）：
  正确 HIT     : 前向期内价格触及目标价（任一），或最高收益 ≥ 目标幅度的 60%
  错误 MISS    : 前向期内价格触止损，且未触及目标
  无效 NEUTRAL : 前向期内价格波动 < 2%，既不到目标也不到止损
  待定 PENDING : 信号日期太近，前向数据不足

跟踪周期：5 / 10 / 20 个交易日

用法：
  python3 modules/tracker.py                        # 跟踪所有历史信号
  python3 modules/tracker.py --date 2026-05-29      # 指定截止日期
  python3 modules/tracker.py add 600519 贵州茅台 --reason "消费龙头反弹"  # 新增标的
"""
from __future__ import annotations

import json
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

# Ensure parent is on path for standalone execution
_root = Path(__file__).parent.parent
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

from core.trader_mcp import mcp_call

# ── 路径 ────────────────────────────────────────────────
ROOT_DIR   = Path(__file__).parent.parent
OUTPUT_DIR = ROOT_DIR / "output" / "tracker"
DATA_DIR   = ROOT_DIR / "data"
PANEL_PATH = Path("/Users/scott/Desktop/一小步/stock panel.xlsx")
SIGNAL_LOG = DATA_DIR / "strategy_log.jsonl"
TRACK_LOG  = DATA_DIR / "tracker_log.jsonl"

HORIZONS = [5, 10, 20]


# ══════════════════════════════════════════════════════════════
# 1. 数据加载
# ══════════════════════════════════════════════════════════════

def _load_signals() -> list[dict]:
    """从 strategy_log.jsonl 读取所有历史信号。"""
    if not SIGNAL_LOG.exists():
        return []
    signals = []
    with open(SIGNAL_LOG, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                signals.append(json.loads(line))
    return signals


def _load_existing_tracks() -> dict[str, dict]:
    """加载已有的跟踪记录，按 (date, code, horizon) 为 key。"""
    if not TRACK_LOG.exists():
        return {}
    tracks = {}
    with open(TRACK_LOG, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                r = json.loads(line)
                key = f"{r['signal_date']}|{r['code']}|{r['horizon']}"
                tracks[key] = r
    return tracks


# ══════════════════════════════════════════════════════════════
# 2. 前向 K 线获取
# ══════════════════════════════════════════════════════════════

def _forward_kline(code: str, start_date: str, days: int) -> list[dict]:
    """
    获取 start_date 之后最多 days 个交易日的日线。
    返回 [{date, open, high, low, close}] 列表，按日期升序。
    """
    end = _add_trade_days(start_date, days + 10)
    result = mcp_call("market_quote", "get_kline", {
        "keyword":           code,
        "start_date":        start_date,
        "end_date":          end,
        "kline_type":        1,
        "reinstatement_type": 2,
    })

    raw = result if isinstance(result, list) else result.get("list", [])
    bars = []
    for b in raw:
        date_str = b.get("trade_date", "")
        if date_str <= start_date:
            continue
        bars.append({
            "date":  date_str,
            "open":  float(b.get("open_price", 0)),
            "high":  float(b.get("high_price", 0)),
            "low":   float(b.get("low_price", 0)),
            "close": float(b.get("close_price", 0)),
        })

    bars.sort(key=lambda b: b["date"])
    return bars[:days]


def _add_trade_days(date_str: str, n: int) -> str:
    """粗略加 n 个日历日（实际交易日可能少一些，K线接口会处理）。"""
    from datetime import datetime as dt
    d = dt.strptime(date_str, "%Y-%m-%d")
    return (d + timedelta(days=int(n * 1.4))).strftime("%Y-%m-%d")


# ══════════════════════════════════════════════════════════════
# 3. 核心判定引擎（自建交易逻辑）
# ══════════════════════════════════════════════════════════════

def _judge_one(
    code: str,
    signal_date: str,
    entry: float,
    stop: float,
    targets: list[float],
    horizon: int,
) -> dict:
    """
    判断一个信号在给定前向窗口内的绩效。

    返回:
      result      : HIT | MISS | NEUTRAL | NODATA | PENDING
      max_return  : 最高价/入场 - 1 (%)
      max_dd      : 最低价/入场 - 1 (%, 负值)
      final_return: 收盘价/入场 - 1 (%)
      hit_target  : 触及的目标价（None 若未触）
      hit_stop    : 是否触及止损
      days_to_target: 触及目标所需天数
      days_to_stop  : 触及止损所需天数
      n_bars        : 实际获取的 K 线条数
    """
    base = {
        "code": code,
        "signal_date": signal_date,
        "horizon": horizon,
        "entry": round(entry, 2),
        "stop": round(stop, 2),
        "targets": [round(t, 2) for t in targets],
        "result": "PENDING",
        "max_return": 0.0,
        "max_dd": 0.0,
        "final_return": 0.0,
        "hit_target": None,
        "hit_stop": False,
        "days_to_target": None,
        "days_to_stop": None,
        "n_bars": 0,
        "track_date": datetime.now().strftime("%Y-%m-%d"),
    }

    if entry <= 0 or stop <= 0:
        base["result"] = "NODATA"
        return base

    # 获取前向 K 线
    bars = _forward_kline(code, signal_date, horizon)

    if not bars:
        base["result"] = "NODATA"
        return base

    # 检查是否还在 PENDING（数据不够 horizon 天，且信号日期较近）
    n = len(bars)
    base["n_bars"] = n

    if n < 3:
        # 太少，无法判断
        sig_dt = datetime.strptime(signal_date, "%Y-%m-%d")
        if (datetime.now() - sig_dt).days < horizon * 1.4:
            base["result"] = "PENDING"
        else:
            base["result"] = "NODATA"
        return base

    # ── 逐日扫描 ──
    hit_target_on = None
    hit_stop_on   = None
    max_price = entry
    min_price = entry

    for i, bar in enumerate(bars):
        high = float(bar.get("high", 0))
        low  = float(bar.get("low", 0))
        if high <= 0 or low <= 0:
            continue

        max_price = max(max_price, high)
        min_price = min(min_price, low)

        # 检查止损
        if hit_stop_on is None and low <= stop:
            hit_stop_on = i + 1  # 1-indexed 天数

        # 检查目标（任一目标价）
        if hit_target_on is None:
            for t in targets:
                if t > 0 and high >= t:
                    hit_target_on = i + 1
                    break

    # 收盘价
    final_close = float(bars[-1].get("close", entry))

    # 计算收益率
    max_ret  = round((max_price / entry - 1) * 100, 2)
    max_dd   = round((min_price / entry - 1) * 100, 2)
    final_ret = round((final_close / entry - 1) * 100, 2)

    # ── 判定结果 ──
    if hit_target_on is not None and hit_stop_on is not None:
        # 两者都触发：看谁先
        if hit_target_on <= hit_stop_on:
            result = "HIT"
        else:
            result = "MISS"
    elif hit_target_on is not None:
        result = "HIT"
    elif hit_stop_on is not None:
        result = "MISS"
    else:
        # 都没触发：判断动量
        if max_ret >= 2.0:
            # 虽然没有精准触目标，但方向对了且幅度尚可
            result = "HIT_SOFT"  # 方向正确但未达目标
        elif max_dd <= -3.0:
            result = "MISS"
        elif abs(final_ret) < 2.0:
            result = "NEUTRAL"
        elif final_ret > 0:
            result = "HIT_SOFT"
        else:
            result = "MISS"

    base.update({
        "result": result,
        "max_return": max_ret,
        "max_dd": max_dd,
        "final_return": final_ret,
        "hit_target": hit_target_on is not None,
        "hit_stop": hit_stop_on is not None,
        "days_to_target": hit_target_on,
        "days_to_stop": hit_stop_on,
        "n_bars": n,
    })
    return base


# ══════════════════════════════════════════════════════════════
# 4. 批量跟踪主流程
# ══════════════════════════════════════════════════════════════

def _signal_is_actionable(s: dict) -> bool:
    """信号是否值得跟踪（有入场价 + 非异常 + 非观望）。"""
    if s.get("error"):
        return False
    if s.get("signal_type", "—") == "—":
        return False
    if s.get("entry_low", 0) <= 0:
        return False
    return True


def run_track(date: str | None = None, verbose: bool = True) -> list[dict]:
    """
    跟踪所有历史信号的前向绩效。

    对每个有信号的历史记录，计算 5/10/20 日前向表现。
    已存在的记录跳过（不重复计算）。
    """
    signals = _load_signals()
    existing = _load_existing_tracks()

    if not signals:
        if verbose:
            print("[tracker] 无历史信号数据")
        return []

    # 过滤：只跟踪有入场价且非观望的信号
    actionable = [s for s in signals if _signal_is_actionable(s)]

    if verbose:
        print(f"[tracker] 历史信号 {len(signals)} 条，可跟踪 {len(actionable)} 条")

    new_tracks = []
    skipped = 0

    for s in actionable:
        code = s["code"]
        name = s.get("name", "")
        sig_date = s["date"]
        entry = s["entry_low"]  # 入场区间下沿
        stop  = s["stop_loss"]
        targets = s.get("take_profit", []) or []

        if not targets:
            # 无目标价，用 entry * 1.05 作为默认目标
            targets = [round(entry * 1.05, 2)]

        for h in HORIZONS:
            key = f"{sig_date}|{code}|{h}"
            if key in existing:
                skipped += 1
                continue

            if verbose:
                print(f"  [{sig_date}] {code} {name} @ {h}d ...", end=" ", flush=True)

            track = _judge_one(code, sig_date, entry, stop, targets, h)
            track["name"] = name
            track["signal_type"] = s.get("signal_type", "")
            track["strategy"] = s.get("strategy", "")
            track["score"] = s.get("score", 0)
            new_tracks.append(track)

            if verbose:
                print(f"→ {track['result']} (max +{track['max_return']}% / dd {track['max_dd']}%)")

    # 保存新记录
    if new_tracks:
        TRACK_LOG.parent.mkdir(parents=True, exist_ok=True)
        with open(TRACK_LOG, "a", encoding="utf-8") as f:
            for t in new_tracks:
                f.write(json.dumps(t, ensure_ascii=False) + "\n")
        if verbose:
            print(f"[tracker] 新增 {len(new_tracks)} 条跟踪记录（跳过 {skipped} 条已有）")
    else:
        if verbose:
            print(f"[tracker] 全部 {skipped} 条已有记录，无需更新")

    # 合并已有 + 新增，生成 HTML
    all_tracks = list(existing.values()) + new_tracks
    if all_tracks:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        today = date or datetime.now().strftime("%Y-%m-%d")
        html = _html_tracker(all_tracks, today)
        html_path = OUTPUT_DIR / f"tracker_{today}.html"
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html)
        latest_path = OUTPUT_DIR / "latest.html"
        with open(latest_path, "w", encoding="utf-8") as f:
            f.write(html)
        if verbose:
            print(f"[tracker] HTML → {html_path}")

    return all_tracks


# ══════════════════════════════════════════════════════════════
# 5. HTML 反思页面
# ══════════════════════════════════════════════════════════════

def _html_tracker(tracks: list[dict], date: str) -> str:
    """生成信号跟踪反思 HTML。"""
    from html import escape

    # ── 聚合统计 ──
    def histats(records: list[dict]) -> dict:
        total = len(records)
        hit  = sum(1 for r in records if r["result"] in ("HIT", "HIT_SOFT"))
        miss = sum(1 for r in records if r["result"] == "MISS")
        neut = sum(1 for r in records if r["result"] == "NEUTRAL")
        pend = sum(1 for r in records if r["result"] == "PENDING")
        nod  = sum(1 for r in records if r["result"] == "NODATA")
        avg_ret = round(sum(r["max_return"] for r in records) / total, 2) if total else 0
        avg_dd  = round(sum(r["max_dd"] for r in records) / total, 2) if total else 0
        hit_rate = round(hit / (hit + miss) * 100, 1) if (hit + miss) > 0 else 0
        return {
            "total": total, "hit": hit, "miss": miss, "neutral": neut,
            "pending": pend, "nodata": nod,
            "avg_max_return": avg_ret, "avg_max_dd": avg_dd,
            "hit_rate": hit_rate,
        }

    overall   = histats(tracks)
    by_signal = {}
    for stype in ["🔥 进攻", "✅ 买入", "🕐 埋伏"]:
        subset = [r for r in tracks if r.get("signal_type", "").startswith(stype[0])]
        if subset:
            by_signal[stype] = histats(subset)

    by_horizon = {}
    for h in HORIZONS:
        subset = [r for r in tracks if r["horizon"] == h]
        if subset:
            by_horizon[h] = histats(subset)

    # ── 系统研判 ──
    if overall["hit_rate"] >= 60:
        verdict = "✅ 策略有效"
        verdict_detail = f"胜率 {overall['hit_rate']}%，信号具备正向期望收益。当前跟踪结果支持继续按策略执行。"
    elif overall["hit_rate"] >= 40:
        verdict = "⚠️ 边缘有效"
        verdict_detail = f"胜率 {overall['hit_rate']}%，接近随机。需要关注信号质量而非数量，仅做高分标的。"
    else:
        verdict = "❌ 信号失效"
        verdict_detail = f"胜率 {overall['hit_rate']}%，低于随机。策略需要重新校准。建议暂停新开仓，复盘信号逻辑。"

    if overall["total"] < 10:
        verdict_detail += f"  采样量不足（{overall['total']}条），结论仅供参考。"

    # ── 构建 HTML ──
    def statcard(label: str, value, color: str = "#1e293b", sub: str = "") -> str:
        return f"""<div class="stat">
  <div class="num" style="color:{color}">{value}</div>
  <div class="label">{label}</div>
  {f'<div class="sub">{sub}</div>' if sub else ''}
</div>"""

    def histable(title: str, stats: dict) -> str:
        hr = stats.get("hit_rate", 0)
        return f"""<table style="margin-bottom:12px">
<tr><th colspan="3" style="text-align:left;font-size:13px;background:#f1f5f9">{title}</th></tr>
<tr>
  <td style="width:60px;text-align:center"><b>{stats['total']}</b><br><span style="font-size:10px;color:#6b7280">总计</span></td>
  <td style="width:60px;text-align:center"><b style="color:#22c55e">{stats['hit']}</b><br><span style="font-size:10px;color:#6b7280">正确</span></td>
  <td style="width:60px;text-align:center"><b style="color:#ef4444">{stats['miss']}</b><br><span style="font-size:10px;color:#6b7280">错误</span></td>
  <td style="width:60px;text-align:center"><b style="color:#6b7280">{stats['neutral']}</b><br><span style="font-size:10px;color:#6b7280">无效</span></td>
  <td style="width:60px;text-align:center"><b style="color:#f59e0b">{stats['pending']}</b><br><span style="font-size:10px;color:#6b7280">待定</span></td>
  <td style="text-align:center"><b style="font-size:16px;color:{'#22c55e' if hr>=60 else '#f59e0b' if hr>=40 else '#ef4444'}">{hr}%</b><br><span style="font-size:10px;color:#6b7280">胜率</span></td>
  <td style="text-align:center"><b>{stats['avg_max_return']}%</b><br><span style="font-size:10px;color:#6b7280">均最大收益</span></td>
  <td style="text-align:center"><b style="color:#ef4444">{stats['avg_max_dd']}%</b><br><span style="font-size:10px;color:#6b7280">均最大回撤</span></td>
</tr>
</table>"""

    # ── 详细记录表（5/10/20 横轴透视）──
    # 按 (signal_date, code) 分组，每组一行，三个周期作为列
    from collections import defaultdict
    groups: dict[tuple, dict] = defaultdict(dict)
    for r in tracks:
        key = (r["signal_date"], r["code"])
        groups[key][r["horizon"]] = r
        groups[key]["_name"] = r.get("name", "")
        groups[key]["_signal_type"] = r.get("signal_type", "")
        groups[key]["_score"] = r.get("score", 0)

    # 排序：信号日降序，分数降序
    grouped = sorted(groups.items(), key=lambda kv: (kv[0][0], -kv[1].get("_score", 0)), reverse=True)

    def _hcell(rec: dict | None, field: str, fmt: str = "s") -> str:
        """Horizon cell renderer with color."""
        if rec is None:
            return '<td style="padding:3px 5px;text-align:center;color:#d1d5db;font-size:10px">—</td>'
        if field == "result":
            v = rec.get("result", "")
            if v in ("HIT", "HIT_SOFT"):
                return '<td style="padding:3px 5px;text-align:center"><span style="color:#22c55e;font-size:11px;font-weight:600">✅</span></td>'
            if v == "MISS":
                return '<td style="padding:3px 5px;text-align:center"><span style="color:#ef4444;font-size:11px;font-weight:600">❌</span></td>'
            if v == "NEUTRAL":
                return '<td style="padding:3px 5px;text-align:center"><span style="color:#6b7280;font-size:11px">—</span></td>'
            if v == "PENDING":
                return '<td style="padding:3px 5px;text-align:center"><span style="color:#f59e0b;font-size:11px">⏳</span></td>'
            return '<td style="padding:3px 5px;text-align:center;color:#9ca3af;font-size:10px">N/A</td>'
        v = rec.get(field, 0)
        if field == "max_return":
            c = "#22c55e" if v > 0 else "#ef4444"
            return f'<td style="padding:3px 5px;text-align:right;font-size:11px;color:{c}">{v:+.1f}%</td>' if rec["result"] != "NODATA" else '<td style="padding:3px 5px;text-align:center;color:#d1d5db;font-size:10px">—</td>'
        if field == "final_return":
            c = "#22c55e" if v > 0 else "#6b7280" if v == 0 else "#ef4444"
            return f'<td style="padding:3px 5px;text-align:right;font-size:11px;color:{c}">{v:+.1f}%</td>' if rec["result"] != "NODATA" else '<td style="padding:3px 5px;text-align:center;color:#d1d5db;font-size:10px">—</td>'
        return f'<td style="padding:3px 5px;text-align:right;font-size:11px">{v}</td>'

    detail_rows = []
    for (sdate, code), hdict in grouped:
        name = hdict.get("_name", "")
        sig = hdict.get("_signal_type", "")
        detail_rows.append(f"""<tr>
  <td style="padding:4px 6px;font-size:11px;white-space:nowrap">{sdate}</td>
  <td style="padding:4px 6px;font-size:11px"><code style="color:#64748b">{code}</code></td>
  <td style="padding:4px 6px;font-size:11px">{escape(name)}</td>
  <td style="padding:4px 6px;font-size:11px;text-align:center">{sig}</td>
  {_hcell(hdict.get(5), 'result')}
  {_hcell(hdict.get(5), 'max_return')}
  {_hcell(hdict.get(5), 'final_return')}
  {_hcell(hdict.get(10), 'result')}
  {_hcell(hdict.get(10), 'max_return')}
  {_hcell(hdict.get(10), 'final_return')}
  {_hcell(hdict.get(20), 'result')}
  {_hcell(hdict.get(20), 'max_return')}
  {_hcell(hdict.get(20), 'final_return')}
</tr>""")

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>信号跟踪 · {date}</title>
<style>
  body {{ font-family: -apple-system, "Microsoft YaHei", sans-serif; margin:0; padding:20px; background:#fafbfc; color:#1e293b; }}
  .container {{ max-width:1400px; margin:0 auto; }}
  h1 {{ font-size:20px; margin:0 0 4px; }}
  .subtitle {{ color:#6b7280; font-size:13px; margin-bottom:16px; }}
  .verdict {{ padding:16px 20px; border-radius:8px; margin-bottom:20px; }}
  .verdict h2 {{ margin:0 0 6px; font-size:16px; }}
  .verdict p {{ margin:0; font-size:13px; line-height:1.6; }}
  .summary {{ display:flex; gap:20px; margin-bottom:20px; flex-wrap:wrap; }}
  .stat {{ background:#fff; border-radius:8px; padding:10px 16px; box-shadow:0 1px 3px rgba(0,0,0,.06); min-width:80px; text-align:center; }}
  .stat .num {{ font-size:26px; font-weight:700; }}
  .stat .label {{ font-size:11px; color:#6b7280; margin-top:2px; }}
  .stat .sub {{ font-size:10px; color:#9ca3af; }}
  table {{ width:100%; border-collapse:collapse; background:#fff; border-radius:8px; overflow:hidden; box-shadow:0 1px 3px rgba(0,0,0,.06); margin-bottom:20px; }}
  th {{ padding:8px 6px; text-align:left; background:#f8fafc; font-size:11px; color:#6b7280; font-weight:600; }}
  td {{ font-size:11px; border-top:1px solid #f1f5f9; }}
  tr:hover {{ background:#f8fafc; }}
  .section-title {{ font-size:14px; font-weight:600; margin:20px 0 10px; padding-bottom:6px; border-bottom:2px solid #e2e8f0; }}
  .disclaimer {{ margin-top:20px; font-size:11px; color:#9ca3af; }}
</style>
</head>
<body>
<div class="container">
<h1>📡 信号跟踪 · 反思</h1>
<div class="subtitle">{date} · 跟踪 {overall['total']} 条信号记录 · 覆盖 {HORIZONS[0]}/{HORIZONS[1]}/{HORIZONS[2]} 日前向窗口</div>

<div class="verdict" style="background:{'#f0fdf4' if overall['hit_rate']>=60 else '#fefce8' if overall['hit_rate']>=40 else '#fef2f2'}">
  <h2 style="color:{'#22c55e' if overall['hit_rate']>=60 else '#f59e0b' if overall['hit_rate']>=40 else '#ef4444'}">{verdict}</h2>
  <p>{verdict_detail}</p>
</div>

<div class="summary">
  {statcard("总信号", overall['total'], "#1e293b")}
  {statcard("✅ 正确", overall['hit'], "#22c55e")}
  {statcard("❌ 错误", overall['miss'], "#ef4444")}
  {statcard("— 无效", overall['neutral'], "#6b7280")}
  {statcard("⏳ 待定", overall['pending'], "#f59e0b")}
  {statcard("胜率", f"{overall['hit_rate']}%", "#22c55e" if overall['hit_rate']>=60 else "#f59e0b")}
  {statcard("均最大收益", f"{overall['avg_max_return']}%", "#22c55e" if overall['avg_max_return']>0 else "#ef4444")}
  {statcard("均最大回撤", f"{overall['avg_max_dd']}%", "#ef4444")}
</div>

<div class="section-title">按信号类型</div>
{"".join(histable(k, v) for k, v in by_signal.items())}

<div class="section-title">按跟踪周期</div>
{"".join(histable(f"{h}日前向", v) for h, v in by_horizon.items())}

<div class="section-title">详细跟踪记录</div>
<div style="max-height:600px;overflow-y:auto">
<table>
<tr>
  <th style="width:78px">信号日</th>
  <th style="width:68px">代码</th>
  <th style="width:72px">名称</th>
  <th style="width:52px;text-align:center">信号</th>
  <th colspan="3" style="text-align:center;background:#f0fdf4;border-left:2px solid #86efac">← 5日 →</th>
  <th colspan="3" style="text-align:center;background:#eff6ff;border-left:2px solid #93c5fd">← 10日 →</th>
  <th colspan="3" style="text-align:center;background:#fef3c7;border-left:2px solid #fcd34d">← 20日 →</th>
</tr>
<tr>
  <th></th><th></th><th></th><th></th>
  <th style="background:#f0fdf4;width:32px;text-align:center;font-size:10px">判</th>
  <th style="background:#f0fdf4;width:48px;text-align:right;font-size:10px">峰+%</th>
  <th style="background:#f0fdf4;width:48px;text-align:right;font-size:10px">终%</th>
  <th style="background:#eff6ff;border-left:2px solid #93c5fd;width:32px;text-align:center;font-size:10px">判</th>
  <th style="background:#eff6ff;width:48px;text-align:right;font-size:10px">峰+%</th>
  <th style="background:#eff6ff;width:48px;text-align:right;font-size:10px">终%</th>
  <th style="background:#fef3c7;border-left:2px solid #fcd34d;width:32px;text-align:center;font-size:10px">判</th>
  <th style="background:#fef3c7;width:48px;text-align:right;font-size:10px">峰+%</th>
  <th style="background:#fef3c7;width:48px;text-align:right;font-size:10px">终%</th>
</tr>
{"".join(detail_rows)}
</table>
</div>

<div class="disclaimer">
⚠️ 跟踪为事后验证，不代表未来收益。判定基于自建交易逻辑，仅用于信号质量评估。
</div>
</div>
</body>
</html>"""


# ══════════════════════════════════════════════════════════════
# 6. 新增标的
# ══════════════════════════════════════════════════════════════

def add_to_panel(code: str, name: str, reason: str = "") -> bool:
    """
    新增标的到自选池 Excel。

    追加到 stock panel.xlsx 末尾，自动分配序号。
    """
    path = PANEL_PATH
    if not path.exists():
        print(f"❌ 自选池文件不存在: {path}")
        return False

    wb = openpyxl.load_workbook(str(path))
    ws = wb.active

    # 找最大序号
    max_seq = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        seq = row[3] if len(row) > 3 else None
        if isinstance(seq, (int, float)):
            max_seq = max(max_seq, int(seq))

    # 检查是否已存在
    for row in ws.iter_rows(min_row=2, values_only=True):
        existing_code = row[4] if len(row) > 4 else None
        if existing_code is not None and str(existing_code).strip() == code:
            code_s = str(existing_code).strip()
            # 补齐前导零比较
            if code_s.zfill(6) == code.zfill(6):
                print(f"⚠️  {code} 已在自选池中（第 {row[3]} 位）")
                return False

    new_seq = max_seq + 1
    new_row = ws.max_row + 1

    # A=1, B=2, C=3, D=4(序号), E=5(代码), F=6(名称)
    ws.cell(row=new_row, column=4, value=new_seq)
    ws.cell(row=new_row, column=5, value=code)
    ws.cell(row=new_row, column=6, value=name)

    wb.save(str(path))
    print(f"✅ 已添加 {code} {name} → 自选池第 {new_seq} 位" + (f"（{reason}）" if reason else ""))
    return True


# ══════════════════════════════════════════════════════════════
# 7. CLI
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="信号跟踪 + 标的新增")
    sub = parser.add_subparsers(dest="cmd")

    # track
    p_track = sub.add_parser("track", help="运行信号跟踪")
    p_track.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"))

    # add
    p_add = sub.add_parser("add", help="新增标的到自选池")
    p_add.add_argument("code", help="股票代码")
    p_add.add_argument("name", help="股票名称")
    p_add.add_argument("--reason", default="", help="加入理由")

    args = parser.parse_args()

    if args.cmd == "add":
        add_to_panel(args.code, args.name, args.reason)
    elif args.cmd == "track":
        run_track(args.date)
    else:
        # 默认 = track + 自动日期
        run_track(datetime.now().strftime("%Y-%m-%d"))
